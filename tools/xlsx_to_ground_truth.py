#!/usr/bin/env python3
"""
Ground Truth Normalizer: Convert fragmented Excel exports to canonical JSON.

Outputs records in Agent A's SmartrouteRecord-compatible format.

Usage:
    python tools/xlsx_to_ground_truth.py --input data/sample_inspections.xlsx --output data/ground_truth.json
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

# ── Date Normalization ───────────────────────────────────────────────────────

_DATE_FORMATS = [
    "%Y-%m-%d",
    "%m/%d/%Y",
    "%m/%d/%y",
    "%d-%b-%Y",
    "%d-%b-%y",
    "%B %d, %Y",
    "%B %d %Y",
    "%b %d, %Y",
    "%m-%d-%Y",
    "%Y/%m/%d",
]


def normalize_date(raw: str | datetime | None) -> str | None:
    """Convert any date string or datetime object to YYYY-MM-DD."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    raw_str = str(raw).strip()
    if not raw_str:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return raw_str  # best-effort pass-through


# Status mapping compatible with Agent A enums
STATUS_MAP: dict[str, str] = {
    "pass": "pass", "passed": "pass", "approved": "pass", "ok": "pass",
    "p": "pass", "complete": "pass",
    "fail": "fail", "failed": "fail", "rejected": "fail", "denied": "fail",
    "f": "fail",
    "partial": "partial", "conditional": "partial", "cond": "partial",
    "pending": "unknown", "pend": "unknown", "in review": "unknown",
    "under review": "unknown",
}

STRUCTURE_MAP: dict[str, str] = {
    "res": "residential", "residential": "residential", "r": "residential",
    "sfr": "residential", "single family": "residential",
    "comm": "commercial", "commercial": "commercial", "c": "commercial",
    "com": "commercial",
    "ind": "industrial", "industrial": "industrial", "i": "industrial",
    "mixed": "mixed-use", "mixed-use": "mixed-use", "mixed use": "mixed-use",
    "mu": "mixed-use",
}


def normalize_address(street: str | None, city: str | None, zip_code: str | None) -> str:
    """Concatenate fragmented address fields."""
    parts = []
    for p in (street, city, zip_code):
        if p is not None:
            s = str(p).strip()
            if s:
                parts.append(s)
    return ", ".join(parts) if parts else ""


# ── Column Alias Mapping ────────────────────────────────────────────────────

_COLUMN_ALIASES = {
    "permit": "permit_number", "permit_number": "permit_number",
    "permit #": "permit_number", "permit_no": "permit_number", "prmt": "permit_number",
    "date": "release_date", "release_date": "release_date", "release date": "release_date",
    "inspection_date": "release_date", "inspection date": "release_date", "insp date": "release_date",
    "street": "street", "address": "street", "site_address": "street", "site address": "street",
    "city": "city", "zip": "zip", "zip_code": "zip", "zipcode": "zip",
    "type": "structure_type", "structure_type": "structure_type",
    "structure type": "structure_type", "bldg type": "structure_type", "building type": "structure_type",
    "result": "status", "inspection_result": "status", "inspection result": "status", "status": "status",
    "inspector": "inspector_name", "inspector_name": "inspector_name",
    "inspector name": "inspector_name", "inspected by": "inspector_name",
    "notes": "notes", "comments": "notes", "remarks": "notes",
}


def _map_columns(headers: list[str | None]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        key = str(h).strip().lower()
        if key in _COLUMN_ALIASES:
            mapping[i] = _COLUMN_ALIASES[key]
    return mapping


def convert_xlsx(input_path: Path) -> list[dict]:
    """Read an XLSX and return a list of ground-truth record dicts."""
    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Workbook has no active sheet")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Workbook must have at least a header row and one data row")

    headers = list(rows[0])
    col_map = _map_columns([str(h) if h else None for h in headers])

    records: list[dict] = []
    for row in rows[1:]:
        raw: dict[str, str] = {}
        for col_idx, field_name in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            raw[field_name] = str(val).strip() if val is not None else ""

        permit = raw.get("permit_number", "").strip() or ""
        date = normalize_date(raw.get("release_date"))
        address = normalize_address(raw.get("street"), raw.get("city"), raw.get("zip"))
        if not raw.get("city") and not raw.get("zip") and raw.get("street"):
            address = raw["street"]

        status_raw = raw.get("status", "").strip().lower()
        status = STATUS_MAP.get(status_raw, "unknown")

        structure_raw = raw.get("structure_type", "").strip().lower()
        structure_type = STRUCTURE_MAP.get(structure_raw, "")

        inspector = raw.get("inspector_name") or None
        notes = raw.get("notes") or None

        # Output in Agent A compatible ground-truth format
        record = {
            "permit_number": permit,
            "inspection_date": date,
            "status": status,
            "site_address": address,
            "structure_type": structure_type,
            "inspector_name": inspector if inspector else "",
            "notes": notes if notes else "",
        }
        records.append(record)

    wb.close()
    return records


# ── Sample XLSX Generator ────────────────────────────────────────────────────

_SAMPLE_DATA = [
    ["Permit #", "Inspection Date", "Street", "City", "Zip", "Bldg Type", "Status", "Inspector", "Notes"],
    ["BP-2026-00142", "01/15/2026", "123 Main St", "Springfield", "62704", "Res", "Passed", "John Smith", "All clear"],
    ["BP-2026-00143", "1/16/2026", "456 Oak Ave", "Springfield", "62705", "Comm", "Failed", "Jane Doe", "Wiring issues found"],
    ["CP-2026-00201", "15-Jan-2026", "789 Industrial Pkwy", "Shelbyville", "62706", "Ind", "Approved", "Robert Garcia", ""],
    ["BP-2026-00144", "January 17, 2026", "321 Elm Blvd", "Springfield", "62704", "Mixed Use", "Conditional", "Maria Chen", "Needs follow-up"],
    ["BP-2026-00145", "2026-01-18", "654 Pine Dr", "Capital City", "62707", "R", "Pass", "Tom Wilson", "Quick pass"],
    ["IP-2026-00050", "01/19/2026", "100 Factory Ln", "Shelbyville", "62706", "Industrial", "Rejected", "Sarah Kim", "Safety violations"],
    ["BP-2026-00146", "1-Feb-2026", "222 Birch Ct", "Springfield", "62705", "SFR", "P", "James Lee", ""],
    ["CP-2026-00202", "02/03/2026", "333 Commerce Way", "Capital City", "62707", "C", "Pending", "Lisa Wang", "Awaiting docs"],
    ["BP-2026-00147", "Feb 05, 2026", "444 Maple Rd", "Springfield", "62704", "Residential", "Under Review", "David Park", ""],
    ["MP-2026-00010", "2/6/2026", "555 Center Pl", "Springfield", "62705", "MU", "Passed", "Anna Lopez", "Mixed-use approval"],
    ["BP-2026-00148", "02/07/2026", "666 Sunset Ave", "Shelbyville", "62706", "Res", "Fail", "Chris Brown", "Structural issue"],
    ["BP-2026-00149", "2/8/26", "777 Lake Dr", "Capital City", "62707", "Comm", "Approved", "Michelle Lee", "Commercial OK"],
    ["IP-2026-00051", "02/09/2026", "888 River Rd", "Shelbyville", "62706", "Ind", "Cond", "Kevin Nguyen", "Conditional pass"],
    ["BP-2026-00150", "2026-02-10", "999 Valley St", "Springfield", "62704", "Residential", "Passed", "Diane Cruz", ""],
    ["CP-2026-00203", "February 11, 2026", "111 Business Blvd", "Capital City", "62707", "Commercial", "Failed", "Ryan Taylor", "Code violations"],
    ["BP-2026-00151", "2/12/2026", "131 Cherry Ln", "Springfield", "62705", "Res", "Pass", "Amy Johnson", "No issues"],
    ["BP-2026-00152", "02/13/2026", "242 Walnut Ct", "Shelbyville", "62706", "Res", "Pend", "Eric Martinez", "Inspector busy"],
    ["MP-2026-00011", "02/14/2026", "353 Broadway", "Capital City", "62707", "Mixed-Use", "In Review", "Nicole White", "Multi-tenant"],
    ["IP-2026-00052", "2/15/2026", "464 Industrial Ave", "Shelbyville", "62706", "Industrial", "Denied", "Brian Hall", "Major violations"],
    ["BP-2026-00153", "02/16/2026", "575 Oak Pl", "Springfield", "62704", "Res", "Passed", "Stephanie Adams", "Final inspection"],
]


def generate_sample_xlsx(output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Inspections"
    for row in _SAMPLE_DATA:
        ws.append(row)
    wb.save(output_path)
    print(f"✅ Generated sample XLSX: {output_path}")


# ── CLI ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Normalize XLSX inspection data to ground-truth JSON")
    parser.add_argument("--input", "-i", type=Path, help="Input XLSX file")
    parser.add_argument("--output", "-o", type=Path, default=Path("data/ground_truth.json"))
    parser.add_argument("--generate-sample", action="store_true")
    args = parser.parse_args()

    if args.generate_sample or args.input is None:
        sample_path = Path("data/sample_inspections.xlsx")
        generate_sample_xlsx(sample_path)
        if args.input is None:
            args.input = sample_path

    records = convert_xlsx(args.input)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w") as f:
        json.dump(records, f, indent=2)
    print(f"✅ Normalized {len(records)} records → {args.output}")


if __name__ == "__main__":
    main()
